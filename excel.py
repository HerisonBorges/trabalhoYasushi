import openpyxl as xl
from openpyxl.worksheet.table import Table, TableStyleInfo
import pathlib

def salvarExcel ():
        # Verifica se o arquivo Excel existe. Se não, cria com os cabeçalhos
        ficheiro = pathlib.Path("Produtos.xlsx")

        if ficheiro.exists():
            pass
        else:
            workbook = xl.Workbook()
            folha = workbook.active
            folha['A1'] = "Nome do Produto"
            folha['B1'] = "Código do Produto"
            folha['C1'] = "Validade"
            folha['D1'] = "Fornecedor"
            folha['E1'] = "Categoria"
            folha['F1'] = "Unidade"
            folha['G1'] = "Observações"
            
            
            # Definindo a largura das colunas e deixando em negrito o cabeçalho
            for col in range(1, folha.max_column + 1):
                coluna = xl.utils.get_column_letter(col)
                folha.column_dimensions[coluna].width = 20
                folha.cell(row=1, column=col).font = xl.styles.Font(bold=True)
            
            # Criando a tabela
            tab = Table(displayName="TabelaProdutos", ref=f"A1:G{folha.max_row}")
            
            # Estilo da tabela
            estilo = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            
            tab.tableStyleInfo = estilo
            folha.add_table(tab)
            workbook.save("Produtos.xlsx")